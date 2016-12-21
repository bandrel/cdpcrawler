#!/usr/bin/env python
import os
try:
    from netmiko import ConnectHandler
except ImportError:
    print('[!] Looks like you are missing the paramiko library.  run \'pip install netmiko\'')
    if os.name == 'nt':
        print('[!] *NOTE* you must install the Microsoft Visual C++ Compiler for Python 2.7 ' \
              'before installing netmiko.\r\n' \
              'This can be found at http://www.microsoft.com/en-us/download/details.aspx?id=44266')
import getopt
import sys
import re
import getpass
import socket
from openpyxl import Workbook
import threading
import Queue

class WorkerThread(threading.Thread):
    def __init__(self, queue):
        threading.Thread.__init__(self)
        self.queue = queue

    def run(self):
        while True:
            host = self.queue.get()
            connect_to_device(host)
            self.queue.task_done()


# help message
def helpmsg():
    print('Usage: cdpscanner.py [Options] <IP address or Host(optional)>' \
          '  Note: All options are optional.  User is prompted or defaults are used.' \
          '  -h or --help:  This help screen\n' \
          '  -i or --inputfile: specifies a file containing hosts to connect to.\n' \
          '  -u or --username: specifies a username to use\n' \
          '  -p or --password: Specifies the password to use\n' \
          '  -v or --verbose: Enables verbose output\n' \
          '  -t or --telnet:  Enables fallback to telnet\n' \
          '  -o or --output:  Prints the inventory of all of the devices at the end\n' \
          '  -H or --hosts:  specifies hosts via comma seperated values\n' \
          '  -T or --threads:  specifices the number of threads (defaults to 8)\n'\
          '  -g or --graph:  enables graphing of the network and specifices the output file\n')


def getinfo(username, password, host, commands, mode):
    global seen_before
    all_output = []
    row = []
    if mode == "SSH":
        device = {
            'device_type': 'cisco_ios',
            'ip': host,
            'username': username,
            'password': password,
            'port': 22,  # optional, defaults to 22
            'secret': '',  # optional, defaults to ''
            'verbose': False,  # optional, defaults to False
        }
    elif mode == "Telnet":
        device = {
            'device_type': 'cisco_ios_telnet',
            'ip': host,
            'username': username,
            'password': password,
            'port': 23,  # optional, defaults to 22
            'secret': '',  # optional, defaults to ''
            'verbose': False,  # optional, defaults to False
        }
    print('Connecting to %s with %s' % (host, mode))
    # Create instance of SSHClient object
    net_connect = ConnectHandler(**device)
    # Automatically add untrusted hosts (make sure okay for security policy in your environment)
    print('%s connection established to %s' % (mode, host))
    # Use invoke_shell to establish an 'interactive session'
    a = re.search(r'(.*)[#>]', net_connect.find_prompt())
    hostname = a.group(1)
    seen_before.append(hostname)
    ip_lines = net_connect.send_command('show ip int brief | ex unass')
    for line in ip_lines.split('\n'):
        a = re.search(r'([0-9]+\.[0-9]+\.[0-9]+\.[0-9]+)', line)
        if a is not None:
            seen_before.append(a.group(1))
    for command in commands:
        output = net_connect.send_command(command)
        all_output.append(output)
        if command == 'show inventory':
            lines = output.split('\n\n')
            for line in lines:
                new_line = line.replace('\n', '')
                a = re.search(r'NAME:\ ?(\S+).*PID:\ ?(\S+).*SN:\ ?(\S+)', new_line)
                if a is not None:
                    pid = a.group(2)
                    serial = a.group(3)
                    row.append([hostname, pid, serial])
        if verbose_mode:
            print(output)
    return all_output, row

def validate_host(host):
    try:
        socket.inet_aton(host)
        return True
    except:
        try:
            ip_from_host = socket.gethostbyname(device)
            return True
        except:
            print("%s is not a valid host name or IP address" % device)
            sys.exit(2)

def find_hosts_from_output(output):
    neighbor_list = []
    global queue
    global seen_before
    cdp_regex = re.compile(
        r'Device ID:\ *(\S+)(?:[^\r\n]*\r?\n){1,4}'
        r'[\ ]*IP(?:v4)* [aA]ddress:\ (\S+)\r?\n'
        r'Platform: (?:cisco )*(\S+),[\ ]*Capabilities:.*(Switch)')
    for item in output:
        split_output = ['Device ID:'+ e for e in item.split('Device ID:') if e != '']
        for single_device in split_output:
            matches = re.search(cdp_regex, single_device)
            if matches is not None:
                hostname = matches.group(1).split('.')[0]
                fqdn = matches.group(1)
                ip_address = matches.group(2)
                device_model = matches.group(3)
                if hostname not in seen_before and ip_address not in seen_before and fqdn not in seen_before:
                    queue.put(matches.group(2))
                    seen_before.append(hostname)
                    seen_before.append(fqdn)
                    seen_before.append(ip_address)
                neighbor_list.append([hostname, ip_address, device_model])
    return neighbor_list


def connect_to_device(host):
    global seen_before
    global neighbor_list
    global failed_ssh
    global errors_ws
    global inventory_list
    # remove the host you are going to connect to from the set.
    device_output = ''
    seen_before.append(host)
    try:
        # Try SSH and if that fails try telnet.
        device_output, inventory_rows = getinfo(username, password, host, commands, 'SSH')
        neighbor_rows = find_hosts_from_output(device_output)
        for neighbor in neighbor_rows:
            neighbor.insert(0,unicode(inventory_rows[0][0]))
            neighbor.insert(1,unicode(host))
            neighbor_list.append(neighbor)

            [inventory_list.append(row) for row in inventory_rows if row not in inventory_list]

                # Check output for new hostnames
    except Exception as e:
        print("SSH connection to %s failed" % host)
        print(e)
        failed_ssh.append(host)
        errors_ws.append([host, unicode(e), 'SSH'])
        if telnet_enabled:
            try:
                device_output, inventory_rows = getinfo(username, password, host, commands, 'Telnet')
                neighbor_rows = find_hosts_from_output(device_output)
                for neighbor in neighbor_rows:
                    neighbor.insert(0, unicode(inventory_rows[0][0]))
                    neighbor.insert(1, unicode(host))
                    neighbor_ws.append(neighbor)
                    [inventory_list.append(row) for row in inventory_rows if row not in inventory_list]
            except Exception as e:
                print("telnet connection to %s failed" % host)
                failed_telnet.append(host)
                errors_ws.append([host, unicode(e), 'telnet'])
        else:
            pass


# Declaration of global variables
inputfile = None
host_set = set()
username = ''
password = ''
failed_hosts = set()
verbose_mode = False
telnet_enabled = False
graph_output = ''
seen_before = []
device = []
failed_telnet = []
failed_ssh = []
outputfile = 'output.xlsx'
inventory_list = []
neighbor_list = []
commands = ['show cdp neighbor detail',
            'show inventory']
thread_num = 8
# setup excel workbook for output
wb = Workbook()
inventory_ws = wb.active
inventory_ws.title = u'Inventory'
neighbor_ws = wb.create_sheet("Neighbors")
errors_ws = wb.create_sheet("Errors")
# Add headers to worksheets
inventory_list.append(['Hostname', 'Device Model', 'Serial Number'])
neighbor_ws.append(['Hostname','IP Address', 'Neighbor Hostname', 'Neighbor IP', 'Neighbor Model'])
errors_ws.append(['Hostname', 'Error', 'Protocol'])

# Run CLI parser function to set variables altered from defaults by CLI arguments.
try:
    opts, args = getopt.getopt(sys.argv[1:], "i:u:p:hvtH:o:T:g:",
                               ["input=", "user=", "password=", "verbose", "telnet","hosts=", "output=","threads=",
                                "graph="])
except getopt.GetoptError:
    helpmsg()
    sys.exit(2)
for opt, arg in opts:
    if opt == '-h':
        helpmsg()
        sys.exit()
    elif opt in ('-i', '--input'):
        inputfile = arg
        with open(inputfile, 'rb') as hostfile:
            for line in hostfile:
                device = line.rstrip()
                validate_host(device)
                host_set.add(device)
    elif opt in ('-t', '--telnet'):
        telnet_enabled = True
    elif opt in ('-T', '--threads'):
        thread_num = arg
    elif opt in ('-v', '--verbose'):
        verbose_mode = True
    elif opt in ('-o', '--output'):
        outputfile = arg
    elif opt in ('-u', '--user'):
        username = arg
    elif opt in ('-p', '--password'):
        password = arg
    elif opt in ('-g', '--graph'):
        from gengraph import creategraph
        graph_output = arg
    elif opt in ('-H', '--hosts'):
        for i in arg.split(','):
            host_set.add(i)

# Set list of IP addreses to connect to if the -i input file is not used
if not host_set:
    host_set.add(raw_input("Enter Switch Hostname or IP Address: ").upper())
if not username:
    username = raw_input("Enter Username: ")
if not password:
    password = getpass.getpass()

queue = Queue.Queue()

for i in range(int(thread_num)):
    worker = WorkerThread(queue)
    worker.setDaemon(True)
    worker.start()

for x in host_set:
    queue.put(x)

queue.join()

for host in failed_ssh:
    if host in failed_telnet:
        print('[!] %s failed both ssh and telnet' % host)
inventory_dedupe = []
[inventory_dedupe.append(i) for i in inventory_list if i not in inventory_dedupe]
[inventory_ws.append(row) for row in inventory_list]
neighbor_dedupe = []
[neighbor_dedupe.append(i) for i in neighbor_list if i not in neighbor_dedupe]
[neighbor_ws.append(row) for row in neighbor_dedupe]
wb.save(outputfile)
if graph_output:
    creategraph(outputfile,graph_output)
