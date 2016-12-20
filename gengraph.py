#!/usr/bin/env python
import pydot
import openpyxl
import sys

def creategraph(inputfile,outputfile):
    wb = openpyxl.load_workbook(inputfile, read_only=True)
    ws = wb['Neighbors']
    graph = pydot.Dot(graph_type='graph', layout='fdp')
    nodes = [pydot.Node(i[0].value, style='Filled', fillcolor='gray') for i in ws.iter_rows(min_row=2)]
    [graph.add_node(i) for i in nodes]
    existing_edges = []
    for i in ws.iter_rows(min_row=2):
        if (i[0].value, i[2].value) not in existing_edges:
            existing_edges.append((i[0].value, i[2].value))
            existing_edges.append((i[2].value, i[0].value))
            graph.add_edge(pydot.Edge(i[0].value, i[2].value))
    graph.write_png(outputfile)

if __name__ == "__main__":
    creategraph(sys.argv[1],sys.argv[2])